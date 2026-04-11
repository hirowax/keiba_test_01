#!/usr/bin/env python3
"""
netkeiba タイム指数スクレイパー
使い方: python scraper.py [YYYYMMDD]
"""

import sys
import json
import time
import random
import logging
import re
from pathlib import Path
from datetime import datetime
from typing import Optional, Dict, List

from dotenv import load_dotenv
import os
import pandas as pd
from openpyxl.styles import PatternFill
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
from bs4 import BeautifulSoup

YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# ─── ログ設定 ────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler()],
)
logger = logging.getLogger(__name__)

# ─── 定数 ────────────────────────────────────────────────────
COOKIES_FILE = Path("cookies.json")
LOGIN_URL = "https://regist.netkeiba.com/account/?pid=login"
RACE_LIST_URL = "https://race.netkeiba.com/top/race_list.html?kaisai_date={date}"
SPEED_URL      = "https://race.netkeiba.com/race/speed.html?race_id={race_id}&type=rank&mode={mode}"
SPEED_URL_BASE = "https://race.netkeiba.com/race/speed.html?race_id={race_id}&rf=shutuba_submenu"
MODES = {
    "average": "近走平均",
    "distance": "当該距離",
    "course": "当該コース",
}
ACCESS_INTERVAL = 3  # 秒（後方互換のため残す）

# 中間経由候補ページ
_BROWSE_VIA = [
    "https://www.netkeiba.com/",
    "https://race.netkeiba.com/",
    "https://db.netkeiba.com/",
    "https://news.netkeiba.com/",
    "https://race.netkeiba.com/top/news.html",
]


def human_sleep(min_sec: float = 3.0, max_sec: float = 8.0) -> None:
    """人間らしいランダム待機"""
    t = random.uniform(min_sec, max_sec)
    # まれに少し長い「考え中」ポーズを入れる
    if random.random() < 0.12:
        t += random.uniform(2.0, 6.0)
    time.sleep(t)


def _random_scroll(page) -> None:
    """ページをランダムにスクロール（人間らしい読み込み動作）"""
    try:
        for _ in range(random.randint(1, 3)):
            page.evaluate(f"window.scrollBy(0, {random.randint(80, 500)})")
            time.sleep(random.uniform(0.2, 0.9))
        if random.random() < 0.45:
            page.evaluate(f"window.scrollBy(0, -{random.randint(40, 200)})")
            time.sleep(random.uniform(0.1, 0.5))
    except Exception:
        pass


def human_browse(page, target_url: str, force_via: str = None) -> None:
    """
    ランダムな中間ページを経由してから target_url へ移動。
    直接ジャンプを避け、人間の自然な閲覧に見せる。
    force_via: 指定した場合は必ずそのページを経由する
    """
    # 約35%の確率で中間ページ経由（force_via指定時は必ず経由）
    if force_via or random.random() < 0.35:
        via = force_via or random.choice(_BROWSE_VIA)
        try:
            page.goto(via, wait_until="domcontentloaded")
            _random_scroll(page)
            human_sleep(1.2, 4.0)
        except Exception:
            pass

    page.goto(target_url, wait_until="domcontentloaded")


# ─── ユーティリティ ───────────────────────────────────────────
def load_env() -> tuple[str, str]:
    load_dotenv()
    email = os.getenv("NETKEIBA_EMAIL")
    password = os.getenv("NETKEIBA_PASSWORD")
    if not email or not password:
        raise ValueError(".env に NETKEIBA_EMAIL / NETKEIBA_PASSWORD が設定されていません")
    return email, password


def save_cookies(context) -> None:
    cookies = context.cookies()
    COOKIES_FILE.write_text(json.dumps(cookies, ensure_ascii=False, indent=2))
    logger.info(f"cookies を保存しました: {COOKIES_FILE}")


def load_cookies(context) -> bool:
    if not COOKIES_FILE.exists():
        return False
    cookies = json.loads(COOKIES_FILE.read_text())
    context.add_cookies(cookies)
    logger.info("既存の cookies を読み込みました")
    return True


def is_logged_in(page) -> bool:
    """ログイン済みかどうかをマイページリンクで確認"""
    page.goto("https://www.netkeiba.com/", wait_until="domcontentloaded")
    return page.locator("a[href*='mypage']").count() > 0 or \
           page.locator("a:has-text('ログアウト')").count() > 0


# ─── ログイン ─────────────────────────────────────────────────
def login(page, email: str, password: str) -> None:
    logger.info("ログインを試みます")
    # まず netkeiba トップを経由してから
    try:
        page.goto("https://www.netkeiba.com/", wait_until="domcontentloaded")
        _random_scroll(page)
        human_sleep(1.0, 3.0)
    except Exception:
        pass
    page.goto(LOGIN_URL, wait_until="domcontentloaded")
    human_sleep(0.8, 2.5)
    page.fill("input[name='login_id']", email)
    time.sleep(random.uniform(0.3, 0.9))
    page.fill("input[name='pswd']", password)
    human_sleep(0.5, 1.5)
    page.click("input[type='submit'], button[type='submit']")
    try:
        page.wait_for_load_state("load", timeout=15000)
    except PlaywrightTimeoutError:
        pass
    human_sleep(1.5, 3.5)
    logger.info("ログイン完了")


# ─── レース一覧取得 ───────────────────────────────────────────
def get_race_ids(page, date: str) -> List[dict]:
    """
    指定日の全レース情報を返す
    Returns: [{"race_id": "...", "label": "中山11R"}, ...]
    """
    url = RACE_LIST_URL.format(date=date)
    logger.info(f"レース一覧取得: {url}")
    # race.netkeiba.com トップを必ず経由
    human_browse(page, url, force_via="https://race.netkeiba.com/")
    _random_scroll(page)

    # レース一覧テーブルが描画されるのを待つ
    try:
        page.wait_for_selector("dl.RaceList_DataList, .RaceList, a[href*='/race/result']", timeout=15000)
    except PlaywrightTimeoutError:
        logger.warning("レース一覧セレクタがタイムアウト。HTMLをそのままパース")

    html = page.content()
    soup = BeautifulSoup(html, "html.parser")

    races = []
    seen = set()

    # race_id を href から抽出
    for a in soup.find_all("a", href=True):
        href = a["href"]
        # /race/result.html?race_id=XXXXXX 形式
        m = re.search(r"race_id=(\d{12})", href)
        if not m:
            continue
        race_id = m.group(1)
        if race_id in seen:
            continue
        seen.add(race_id)

        # ラベル生成（競馬場名 + レース番号）
        label = _build_label(race_id, soup, a)
        races.append({"race_id": race_id, "label": label})

    logger.info(f"{len(races)} レースを検出")
    return races


def _build_label(race_id: str, soup: BeautifulSoup, a_tag) -> str:
    """race_id と周辺HTMLからラベル（例: 中山11R）を生成"""
    # race_id の構造: YYYYCCRRBB  (C=場コード, RR=回, BB=日, ?? = レース番号)
    # 実際: 12桁 = YYYY(4) + 場(2) + 回(1) + 日(1) + レース(2)
    venue_code = race_id[4:6]
    race_num_str = race_id[10:12].lstrip("0") or "1"

    venue_map = {
        "01": "札幌", "02": "函館", "03": "福島", "04": "新潟",
        "05": "東京", "06": "中山", "07": "中京", "08": "京都",
        "09": "阪神", "10": "小倉",
    }
    venue = venue_map.get(venue_code, f"場{venue_code}")
    return f"{venue}{race_num_str}R"


# ─── タイム指数テーブルパース ──────────────────────────────────
def _is_valid_speed_df(df: pd.DataFrame) -> bool:
    """DataFrameが正しい指数ランキング表かチェック（会員プラン表でないか）"""
    if df is None or df.empty:
        return False
    # 先頭列に "N位" 形式の行が存在するかどうか
    first_col = df.iloc[:, 0].astype(str)
    return first_col.str.match(r"^\d+位$").any()


def preflight_premium_check(page, race_id: str) -> bool:
    """プレミアムコンテンツにアクセスできるか検証（起動時1回だけ）。
    type=rank ページが順位データを返すか確認する。"""
    url = SPEED_URL.format(race_id=race_id, mode="average")
    try:
        page.goto(url, wait_until="domcontentloaded")
        time.sleep(random.uniform(3.0, 5.0))
    except Exception:
        return False
    soup = BeautifulSoup(page.content(), "html.parser")
    tables = soup.find_all("table")
    if not tables:
        return False
    target = max(tables, key=lambda t: len(t.find_all("tr")))
    return "1位" in target.get_text()


def parse_speed_table(page, race_id: str, mode: str) -> Optional[pd.DataFrame]:
    """type=rank&mode={mode} でタイム指数ランキング表を取得。
    データ未公開の場合(前日スクレイプ等)は None を返す。"""
    url = SPEED_URL.format(race_id=race_id, mode=mode)
    try:
        if mode == "average" and random.random() < 0.4:
            via = f"https://race.netkeiba.com/race/shutuba.html?race_id={race_id}"
            human_browse(page, url, force_via=via)
        else:
            human_browse(page, url)
        _random_scroll(page)
        page.wait_for_selector("table", timeout=15000)
    except PlaywrightTimeoutError:
        logger.warning(f"テーブル待機タイムアウト: {url}")

    html = page.content()
    soup = BeautifulSoup(html, "html.parser")

    tables = soup.find_all("table")
    if not tables:
        return None

    # 最大行数テーブルを候補とする
    target = max(tables, key=lambda t: len(t.find_all("tr")))
    rows = target.find_all("tr")
    if len(rows) < 2:
        return None

    header_cells = rows[0].find_all(["th", "td"])
    headers = [c.get_text(strip=True) for c in header_cells]
    data = []
    for row in rows[1:]:
        cells = row.find_all(["th", "td"])
        if not cells:
            continue
        data.append([c.get_text(strip=True) for c in cells])
    if not data:
        return None

    max_cols = max(len(headers), max(len(r) for r in data))
    headers += [f"col{i}" for i in range(len(headers), max_cols)]
    for row in data:
        row += [""] * (max_cols - len(row))
    df = pd.DataFrame(data, columns=headers[:max_cols])

    # 会員プラン表などが返ってきた場合は None
    if not _is_valid_speed_df(df):
        logger.warning(f"指数ランキング表ではないコンテンツを取得: {url}")
        return None
    return df


def parse_speed_shutuba(page, race_id: str) -> Optional[Dict[str, pd.DataFrame]]:
    """rf=shutuba_submenu URLから全3指数を一括取得。
    type=rank が未公開(前日等)のときのフォールバック用。
    Returns: {"average": df, "distance": df, "course": df} または None
    各dfはtype=rank形式: 順位列(col0)="N位", 馬番(col2), 馬名(col4), 指数(col5)

    CSSクラスで直接取得:
      sk__umaban          → 馬番
      Horse_Name          → 馬名 (<a>のテキスト)
      sk__average_index   → ５走平均 (Sort_Function_Data_Hidden span)
      sk__max_distance_index → 距離  (同上、"未"を含む場合はデータなし)
      sk__max_course_index   → コ｜ス (同上)
    """
    url = SPEED_URL_BASE.format(race_id=race_id)
    try:
        human_browse(page, url)
        _random_scroll(page)
        page.wait_for_selector("table", timeout=15000)
    except PlaywrightTimeoutError:
        logger.warning(f"shutuba速度表タイムアウト: {url}")
        return None

    html = page.content()
    soup = BeautifulSoup(html, "html.parser")

    def _cell_score(cell) -> str:
        """CSSセルから指数値を抽出。Sort_Function_Data_Hidden spanを優先。
        "未"を含む場合はデータなし → "" を返す。"""
        if cell is None:
            return ""
        full_text = cell.get_text(strip=True)
        if "未" in full_text:
            return ""
        hidden = cell.find(class_="Sort_Function_Data_Hidden")
        if hidden:
            v = hidden.get_text(strip=True)
            try:
                return str(int(v))
            except ValueError:
                pass
        # フォールバック: 先頭の整数を抽出
        m = re.search(r"\d+", full_text)
        return m.group() if m else ""

    def _cell_name(cell) -> str:
        """馬名セルからリンクテキストを取得（Sort_Function_Data_Hidden を除外）"""
        if cell is None:
            return ""
        a = cell.find("a")
        if a:
            return re.sub(r"\s+", "", a.get_text())
        # フォールバック: Sort_Function_Data_Hiddenを除いたテキスト
        for hidden in cell.find_all(class_="Sort_Function_Data_Hidden"):
            hidden.decompose()
        return re.sub(r"\s+", "", cell.get_text())

    # テーブル内のデータ行を走査
    table = soup.find("table")
    if not table:
        logger.warning(f"shutuba: テーブルなし ({url})")
        return None

    horses = []
    for tr in table.find_all("tr"):
        # 馬番セルの有無でデータ行を判定
        num_cell  = tr.find(class_=re.compile(r"sk__umaban|UmaBan"))
        name_cell = tr.find(class_=re.compile(r"Horse_Name|sk__horse_name"))
        if not num_cell or not name_cell:
            continue
        num = num_cell.get_text(strip=True)
        if not num.isdigit():
            continue
        name = _cell_name(name_cell)
        if not name:
            continue

        avg_cell  = tr.find(class_="sk__average_index")
        dist_cell = tr.find(class_="sk__max_distance_index")
        crs_cell  = tr.find(class_="sk__max_course_index")

        horses.append({
            "num":  num,
            "name": name,
            "avg":  _cell_score(avg_cell),
            "dist": _cell_score(dist_cell),
            "crs":  _cell_score(crs_cell),
        })

    if not horses:
        logger.warning(f"shutuba: 馬データ取得失敗 ({url})")
        return None

    def make_rank_df(horses, val_key):
        """指定列でソートしてtype=rank形式のDataFrameを作る"""
        scored = []
        for h in horses:
            v = h[val_key]
            try:
                scored.append((float(v), h))
            except (ValueError, TypeError):
                scored.append((-1.0, h))  # 値なしは末尾
        scored.sort(key=lambda x: -x[0])

        rows_out = []
        rank = 1
        prev_score = None
        rank_counter = 0
        for score, h in scored:
            rank_counter += 1
            if score != prev_score and score > 0:
                rank = rank_counter
            prev_score = score
            rank_str = f"{rank}位" if score > 0 else "-"
            # type=rank形式: [順位, 枠(空), 馬番, 印(空), 馬名, 指数]
            idx_val = str(int(score)) if score > 0 else "-"
            rows_out.append([rank_str, "", h["num"], "", h["name"], idx_val])
        headers = ["順位", "枠", "馬番", "印", "馬名", "指数"]
        return pd.DataFrame(rows_out, columns=headers)

    # 各モードのデータが1頭以上あれば結果に含める
    result = {}
    for key, val_key in [("average", "avg"), ("distance", "dist"), ("course", "crs")]:
        if any(h[val_key] for h in horses):
            result[key] = make_rank_df(horses, val_key)

    if not result:
        logger.warning(f"shutuba: 指数データなし ({url})")
        return None

    logger.info(f"shutuba fallback 成功: {race_id} ({len(horses)}頭, {list(result.keys())})")
    return result


# ─── Excel 書き出し ───────────────────────────────────────────
def venue_of(label: str) -> str:
    """'中山11R' → '中山'"""
    return re.sub(r"\d+R$", "", label)


def write_excel(
    all_data: Dict[str, Dict[str, Dict[str, Optional[pd.DataFrame]]]],
    all_summaries: Dict[str, Dict[str, pd.DataFrame]],
    date: str,
    triple_rows: List[dict],
) -> None:
    """
    all_data:      venue -> race_label -> mode_label -> df
    all_summaries: venue -> race_label -> summary_df
    triple_rows:   3指数重複馬リスト（黄色ハイライト用）
    """
    Path("output").mkdir(exist_ok=True)
    Path("summary").mkdir(exist_ok=True)

    # 古い個別CSVを削除（全場_3指数重複馬.csv は残す）
    old_dir = Path("output") / date
    if old_dir.exists():
        for f in old_dir.glob("*.csv"):
            if f.name != "全場_3指数重複馬.csv":
                f.unlink()

    # 黄色ハイライト対象の (開催場, レース番号, 馬番) セット
    triple_set = set(
        (r["開催場"], r["レース番号"], str(r["馬番"]).strip())
        for r in triple_rows
    )

    output_path = Path("output") / f"{date}.xlsx"
    summary_path = Path("summary") / f"{date}.xlsx"

    # ── タイム指数 Excel ──
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for venue, races in all_data.items():
            row_cursor = 0
            sheet_written = False
            for race_label, mode_dfs in races.items():
                race_num = race_label[len(venue):]  # "中山11R" → "11R"
                for mode_label, df in mode_dfs.items():
                    if df is None:
                        continue
                    # セクションヘッダー行
                    header_df = pd.DataFrame([[f"■ {race_label}  {mode_label}"]])
                    header_df.to_excel(
                        writer, sheet_name=venue,
                        startrow=row_cursor, index=False, header=False
                    )
                    row_cursor += 1
                    df_start = row_cursor
                    df.to_excel(
                        writer, sheet_name=venue,
                        startrow=df_start, index=False
                    )

                    # 馬番セルを黄色ハイライト
                    num_col_idx = next(
                        (i for i, c in enumerate(df.columns) if "馬番" in c), None
                    )
                    if num_col_idx is not None and triple_set:
                        ws = writer.sheets[venue]
                        for row_i, (_, row) in enumerate(df.iterrows()):
                            num_val = str(row.iloc[num_col_idx]).strip()
                            if (venue, race_num, num_val) in triple_set:
                                # openpyxl は1始まり、df_start+1=列ヘッダー行、df_start+2=データ1行目
                                ws.cell(
                                    row=df_start + 2 + row_i,
                                    column=num_col_idx + 1
                                ).fill = YELLOW_FILL

                    row_cursor += len(df) + 2  # データ + 空行
                    sheet_written = True
            if sheet_written:
                logger.info(f"[output] {venue} シート書き込み完了")

    logger.info(f"保存: {output_path}")

    # ── サマリー Excel ──
    with pd.ExcelWriter(summary_path, engine="openpyxl") as writer:
        for venue, races in all_summaries.items():
            row_cursor = 0
            sheet_written = False
            for race_label, summary_df in races.items():
                header_df = pd.DataFrame([[f"■ {race_label}"]])
                header_df.to_excel(
                    writer, sheet_name=venue,
                    startrow=row_cursor, index=False, header=False
                )
                row_cursor += 1
                summary_df.to_excel(
                    writer, sheet_name=venue,
                    startrow=row_cursor, index=False
                )
                row_cursor += len(summary_df) + 2
                sheet_written = True
            if sheet_written:
                logger.info(f"[summary] {venue} シート書き込み完了")

    logger.info(f"保存: {summary_path}")


# ─── サマリー生成 ─────────────────────────────────────────────
def _get_topN(df: pd.DataFrame, n: int) -> Optional[pd.DataFrame]:
    """指数列を特定して上位N頭を返す"""
    idx_col = None
    for col in df.columns:
        if "指数" in col:
            idx_col = col
            break
    if idx_col is None:
        numeric_cols = df.select_dtypes(include="number").columns.tolist()
        if len(numeric_cols) >= 1:
            idx_col = numeric_cols[0]

    if idx_col is None:
        return None

    horse_col = next((c for c in df.columns if "馬名" in c), None)
    num_col = next((c for c in df.columns if "馬番" in c), None)

    df = df.copy()
    df[idx_col] = pd.to_numeric(df[idx_col], errors="coerce")
    df = df.dropna(subset=[idx_col])
    df = df.sort_values(idx_col, ascending=False).head(n)

    cols = [c for c in [num_col, horse_col, idx_col] if c is not None]
    if not cols:
        return None
    return df[cols].reset_index(drop=True)


def _get_top5(df: pd.DataFrame) -> Optional[pd.DataFrame]:
    """指数列を特定してトップ5を返す（Excel表示用）"""
    return _get_topN(df, 5)


def build_summary(dfs: Dict[str, Optional[pd.DataFrame]], label: str) -> pd.DataFrame:
    """
    dfs: {"average": df, "distance": df, "course": df}
    """
    records = []

    # ① 各指数トップ5
    top5_all: Dict[str, Optional[pd.DataFrame]] = {}
    for mode_key, mode_label in MODES.items():
        df = dfs.get(mode_key)
        top5 = _get_top5(df) if df is not None else None
        top5_all[mode_key] = top5

        records.append({"セクション": f"【{mode_label} トップ5】", "馬番": "", "馬名": "", "近走平均": "", "当該距離": "", "当該コース": ""})
        if top5 is not None:
            for _, row in top5.iterrows():
                r = {"セクション": "", "馬番": "", "馬名": "", "近走平均": "", "当該距離": "", "当該コース": ""}
                for col in top5.columns:
                    if "馬番" in col:
                        r["馬番"] = row[col]
                    elif "馬名" in col:
                        r["馬名"] = row[col]
                    elif "指数" in col:
                        r[mode_label] = row[col]
                records.append(r)
        else:
            records.append({"セクション": "データなし", "馬番": "", "馬名": "", "近走平均": "", "当該距離": "", "当該コース": ""})

    # ② 3指数すべてでトップ3に入っている馬
    records.append({"セクション": "【3指数すべてトップ3入り馬】", "馬番": "", "馬名": "", "近走平均": "", "当該距離": "", "当該コース": ""})

    triple_horses = _find_triple_top5(dfs)
    if triple_horses:
        for entry in triple_horses:
            records.append({
                "セクション": "",
                "馬番": entry["馬番"],
                "馬名": entry["馬名"],
                "近走平均": entry.get("近走平均", ""),
                "当該距離": entry.get("当該距離", ""),
                "当該コース": entry.get("当該コース", ""),
            })
    else:
        records.append({"セクション": "該当なし", "馬番": "", "馬名": "", "近走平均": "", "当該距離": "", "当該コース": ""})

    return pd.DataFrame(records)


def find_triple_top5_rows(label: str, dfs: Dict[str, Optional[pd.DataFrame]]) -> List[dict]:
    """
    3指数すべてでトップ3（馬番一致）に入っている馬を返す。
    Returns: [{"開催場", "レース番号", "馬番", "馬名",
               "近走平均指数", "近走平均順位", "当該距離指数", "当該距離順位",
               "当該コース指数", "当該コース順位"}, ...]
    """
    venue = venue_of(label)
    race_num = re.sub(r"^[^\d]+", "", label)  # "中山11R" → "11R"

    # mode_key -> { 馬番: {馬名, 指数, 順位} }
    mode_info: Dict[str, Dict[str, dict]] = {}

    for mode_key, mode_label in MODES.items():
        df = dfs.get(mode_key)
        if df is None:
            return []

        # 列を特定
        idx_col = next((c for c in df.columns if "指数" in c), None)
        num_col = next((c for c in df.columns if "馬番" in c), None)
        horse_col = next((c for c in df.columns if "馬名" in c), None)

        if idx_col is None or num_col is None:
            return []

        work = df.copy()
        work[idx_col] = pd.to_numeric(work[idx_col], errors="coerce")
        work = work.dropna(subset=[idx_col])
        work = work.sort_values(idx_col, ascending=False).reset_index(drop=True)

        table = {}
        for rank_0, row in work.iterrows():
            num = str(row[num_col]).strip()
            if not num or rank_0 >= 3:  # トップ3のみ
                if rank_0 >= 3:
                    break
            table[num] = {
                "馬名": str(row[horse_col]).strip() if horse_col else "",
                "指数": row[idx_col],
                "順位": f"{rank_0 + 1}位",
            }
        mode_info[mode_key] = table

    # 馬番の積集合
    sets = [set(t.keys()) for t in mode_info.values()]
    common_nums = set.intersection(*sets) if sets else set()

    rows = []
    for num in sorted(common_nums, key=lambda x: int(x) if x.isdigit() else 0):
        avg = mode_info["average"][num]
        dist = mode_info["distance"][num]
        crs = mode_info["course"][num]
        rows.append({
            "開催場": venue,
            "レース番号": race_num,
            "馬番": num,
            "馬名": avg["馬名"] or dist["馬名"] or crs["馬名"],
            "近走平均指数": avg["指数"],
            "近走平均順位": avg["順位"],
            "当該距離指数": dist["指数"],
            "当該距離順位": dist["順位"],
            "当該コース指数": crs["指数"],
            "当該コース順位": crs["順位"],
        })
    return rows


def _find_triple_top5(dfs: Dict[str, Optional[pd.DataFrame]]) -> List[dict]:
    """3指数すべてでトップ3入りの馬を返す（サマリー表示用）"""
    top5_names: dict[str, set[str]] = {}
    top5_data: dict[str, dict[str, dict]] = {}  # mode -> {馬名: row_dict}

    for mode_key, mode_label in MODES.items():
        df = dfs.get(mode_key)
        top3 = _get_topN(df, 3) if df is not None else None
        if top3 is None:
            return []

        names = set()
        data = {}
        for _, row in top3.iterrows():
            name = None
            num = None
            idx_val = None
            for col in top3.columns:
                if "馬名" in col:
                    name = str(row[col])
                elif "馬番" in col:
                    num = str(row[col])
                elif "指数" in col:
                    idx_val = row[col]
            if name:
                names.add(name)
                data[name] = {"馬番": num, "馬名": name, mode_label: idx_val}
        top5_names[mode_key] = names
        top5_data[mode_key] = data

    # 積集合
    common = set.intersection(*top5_names.values()) if top5_names else set()
    result = []
    for name in common:
        entry = {"馬番": "", "馬名": name}
        for mode_key, mode_label in MODES.items():
            entry[mode_label] = top5_data[mode_key].get(name, {}).get(mode_label, "")
            if not entry["馬番"] and top5_data[mode_key].get(name, {}).get("馬番"):
                entry["馬番"] = top5_data[mode_key][name]["馬番"]
        result.append(entry)
    return result


# ─── メイン ───────────────────────────────────────────────────
def main():
    # 日付引数
    if len(sys.argv) >= 2:
        date = sys.argv[1]
    else:
        date = datetime.now().strftime("%Y%m%d")
    logger.info(f"対象日: {date}")

    email, password = load_env()

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=True,
            args=["--disable-blink-features=AutomationControlled"],
        )
        context = browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            viewport={"width": 1280, "height": 800},
            locale="ja-JP",
            timezone_id="Asia/Tokyo",
        )
        # headless 検出回避
        context.add_init_script(
            "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
        )
        page = context.new_page()

        # ── cookiesのみ（自動ログインしない → プレミアムcookies保護） ──
        cookie_loaded = load_cookies(context)
        if not cookie_loaded:
            logger.error("❌ cookies.json が見つかりません。")
            logger.error("   → python3 save_cookies.py を実行してログインしてください")
            browser.close()
            return

        # レース一覧取得
        races = get_race_ids(page, date)
        if not races:
            logger.error("レースが見つかりませんでした。終了します。")
            browser.close()
            return

        # ── プリフライトチェック: プレミアムコンテンツにアクセスできるか ──
        test_race_id = races[0]["race_id"]
        if not preflight_premium_check(page, test_race_id):
            logger.error("❌ プレミアムコンテンツにアクセスできません（cookiesが期限切れの可能性）")
            logger.error("   → python3 save_cookies.py を実行して再ログインしてください")
            browser.close()
            return
        logger.info("✅ プリフライトチェック通過（プレミアムアクセス確認済み）")
        human_sleep(5.0, 10.0)

        # venue -> race_label -> mode_label -> df
        all_data: Dict[str, Dict] = {}
        # venue -> race_label -> summary_df
        all_summaries: Dict[str, Dict] = {}
        # 全場 3指数重複馬の行リスト
        triple_rows: List[dict] = []
        fallback_count = 0  # fallback 発動回数（大量発動ガード用）

        for race in races:
            race_id = race["race_id"]
            label = race["label"]
            venue = venue_of(label)
            logger.info(f"処理中: {label} ({race_id})")

            dfs: Dict[str, Optional[pd.DataFrame]] = {}

            for mode_key, mode_label in MODES.items():
                try:
                    df = parse_speed_table(page, race_id, mode_key)
                    dfs[mode_key] = df
                    if df is None:
                        logger.warning(f"{label} {mode_label}: データなし")
                except Exception as e:
                    logger.error(f"{label} {mode_label} 取得失敗: {e}")
                    dfs[mode_key] = None
                finally:
                    human_sleep(3.0, 9.0)

            # フォールバック: type=rank が全て None の場合 shutuba_submenu を試みる
            if all(v is None for v in dfs.values()):
                fallback_count += 1
                logger.warning(f"type=rank 未公開 → shutuba fallback: {race_id}")
                try:
                    fallback = parse_speed_shutuba(page, race_id)
                    if fallback:
                        dfs.update(fallback)
                        got = sum(1 for v in dfs.values() if v is not None)
                        logger.info(f"{label}: shutuba fallback で {got}モード取得")
                except Exception as e:
                    logger.error(f"{label} shutuba fallback 失敗: {e}")
                finally:
                    human_sleep(3.0, 7.0)

                # ── fallback 大量発動ガード ──
                if fallback_count >= 3 and fallback_count >= len(races) * 0.5:
                    logger.error("❌ 過半数のレースで type=rank が取得できません。")
                    logger.error("   cookiesが期限切れか、プレミアム権限に問題があります。")
                    logger.error("   → python3 save_cookies.py を実行して再ログインしてください")
                    browser.close()
                    return

            # mode_label をキーに変換して蓄積
            all_data.setdefault(venue, {})[label] = {
                MODES[k]: v for k, v in dfs.items()
            }

            # サマリー生成
            try:
                summary_df = build_summary(dfs, label)
                all_summaries.setdefault(venue, {})[label] = summary_df
            except Exception as e:
                logger.error(f"{label} サマリー生成失敗: {e}")

            # 3指数重複馬チェック
            try:
                rows = find_triple_top5_rows(label, dfs)
                if rows:
                    logger.info(f"{label}: 3指数重複馬 {len(rows)}頭")
                    triple_rows.extend(rows)
            except Exception as e:
                logger.error(f"{label} 3指数重複馬チェック失敗: {e}")

        browser.close()

        # Excel 一括出力
        write_excel(all_data, all_summaries, date, triple_rows)

        # 全場 3指数重複馬 CSV 出力
        out_dir = Path("output") / date
        out_dir.mkdir(parents=True, exist_ok=True)
        triple_path = out_dir / "全場_3指数重複馬.csv"
        if triple_rows:
            pd.DataFrame(triple_rows).to_csv(triple_path, index=False, encoding="utf-8-sig")
            logger.info(f"保存: {triple_path} ({len(triple_rows)}行)")
        else:
            logger.info("3指数重複馬: 該当レースなし")

        logger.info("完了")


if __name__ == "__main__":
    main()
